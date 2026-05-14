from __future__ import annotations

import asyncio
import json
import math
import re
import shutil
import time
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterable, Optional
from urllib.parse import urlparse

import pandas as pd
import requests
from openpyxl import load_workbook
from PIL import Image, ImageEnhance

from llm_optimizer import LLMConfig, LLMOptimizer
from temu_scraper import ScraperConfig, TemuScraper, specs_to_json


DXM_COLUMNS = [
    "*产品标题",
    "*英文标题",
    "产品描述",
    "产品货号",
    "*变种属性名称一",
    "*变种属性值一",
    "变种属性名称二",
    "变种属性值二",
    "预览图",
    "*申报价格\n(店铺币种)",
    "SKU货号",
    "*长（cm）",
    "*宽（cm）",
    "*高（cm）",
    "*重量（g）",
    "识别码类型",
    "识别码",
    "站外产品链接",
    "*轮播图",
    "*产品素材图",
    "外包装形状",
    "外包装类型",
    "外包装图片",
    "建议售价（USD）",
    "库存",
    "发货时效（天）",
    "是否无属性",
    "变体数量",
]

CLOTHING_KEYWORDS = ("服装", "女装", "男装", "童装", "鞋", "靴", "尺码", "clothing", "shoes", "boots")

DEFAULT_IMAGE_TUNE_PROMPT = """你是一位 Temu 商品轮播图差异化与修图专家。

目标：
在保持原商品完全不变，并保留原图画风的前提下，让图片看起来更干净、更专业，并与普通竞品图片形成一定视觉差异。

请先判断当前图片属于哪一种类型：
1. 纯商品图
2. 使用效果展示图
3. 功能展示图
4. 人物场景图
5. 商品细节特写图
6. 前后对比图
7. 包装或配件图

通用规则：
1. 必须保持原商品完全一致。
2. 保留原图的画风、氛围、真实感、镜头视角和整体视觉语言。
3. 不要改变商品形状、颜色、材质、图案、纹理、数量、包装、配件、文字、Logo 或任何可见商品细节。
4. 不要新增人物、手、图标、标签、水印、促销文字、卖点文字或装饰元素。
5. 不要删除原图中的重要信息，例如功能箭头、步骤、标签、对比结构、商品细节或使用场景。
6. 保留原图用途。
7. 只做轻微视觉差异化，不要让图片看起来像另一个商品，也不要变成另一种摄影风格。
8. 最终图片必须真实、自然，适合 Temu 商品轮播图使用。

允许的差异化优化：
1. 让背景更干净、更简洁、更专业，同时保持相似画风。
2. 在适合的情况下，可以将背景或使用场景替换为相似的真实电商场景。
3. 如果更换场景，必须贴近原图语境、商品品类、光线风格和 Temu 平台风格。
4. 可以适当调整光影方向、明暗层次、阴影柔和度和高光表现，让商品更有立体感。
5. 可以在不改变商品结构和比例的前提下，轻微调整展示角度，让图片与普通竞品图形成差异。
6. 可以对画面进行适当放大或缩小，优化商品在 800x800 画布中的占比。
7. 可以轻微优化构图、留白和视觉重心，但不要裁切商品主体。
8. 自然调整亮度、对比度、锐度和色温。
9. 减少压缩痕迹、噪点、灰尘、杂乱背景和廉价感。
10. 强化真实材质细节，例如纹理、缝线、边缘、厚度和表面质感。
11. 让图片整体更统一、更干净、更适合电商展示。

展示角度和缩放限制：
1. 只允许轻微改变展示角度，不要让商品结构、款式、形状或使用方式发生变化。
2. 不要生成原图中看不到的新背面、新内部结构、新配件或新功能。
3. 放大时不得裁切商品任何部分；缩小时商品仍需清晰可见。
4. 商品在画面中的占比应适合 Temu 轮播图，不能过小，也不能贴边。
5. 光影变化必须真实自然，不要产生塑料感、过度磨皮感或虚假的棚拍效果。

不同图片类型规则：
1. 纯商品图：
   - 可以将背景清理为白色、近白色或简单浅色背景。
   - 可以轻微优化商品居中、留白和自然阴影。
   - 可以轻微调整展示角度、远近比例和商品在画布中的位置。
   - 不要改变商品形状、颜色、图案或真实比例。

2. 使用效果展示图：
   - 保留原本展示的效果、结果和使用含义。
   - 如果有助于差异化，可以替换为相似真实使用场景，但不能改变商品功能。
   - 可以优化光线、背景干净程度和整体画质。
   - 可以轻微调整画面远近和构图，让使用效果更清楚。
   - 不要夸大效果，不要改变效果含义。

3. 功能展示图：
   - 保留所有功能结构、箭头、步骤、标签、零件和可见机制。
   - 可以提升清晰度、对比度和可读性。
   - 不要改写、删除或新增任何文字、图标、箭头或标签。
   - 如果更换场景或调整角度可能破坏功能信息，则不要更换场景或角度。

4. 人物场景图：
   - 保留人物、姿势、体型、肤色、衣着、动作和商品使用方式。
   - 可以轻微清理或替换背景为相似真实生活场景，并保持同一画风。
   - 可以优化光线、背景整洁度、自然色彩和整体高级感。
   - 可以轻微调整画面远近和视觉重心，但不要改变人物动作或商品使用方式。
   - 不要改变人物外貌，不要新增人物或肢体。

5. 商品细节特写图：
   - 保留材质、纹理、缝线、边缘、接口、厚度、图案和细小结构。
   - 可以增强细节清晰度、真实纹理和局部层次。
   - 可以轻微调整局部放大比例，让细节更清楚。
   - 不要过度磨皮、重绘或扭曲材质。
   - 避免大幅更换场景，重点放在干净的细节展示上。

6. 前后对比图：
   - 保留对比结构、边界线、标签、箭头和对比含义。
   - 可以优化亮度、清晰度和阅读体验。
   - 不要改变任一侧展示内容。
   - 避免改变会影响对比结果的场景、角度或比例。

7. 包装或配件图：
   - 保留包装形状、文字、条码、标签、配件数量和摆放方式。
   - 可以优化背景、光线和清晰度。
   - 可以轻微调整画面远近，让包装和配件更清楚。
   - 不要替换包装，不要虚构配件。

最终输出：
只返回修图后的图片，不要输出解释文字。

输出要求：
- 生成 1:1 正方形图片。
- 最终图片尺寸必须为 800 x 800 像素。
- 商品必须完整显示在正方形画布内。
- 不要裁切商品任何部分。
- 商品占比要适中，不能过小，也不能贴边。
- 使用干净的电商构图，适合 Temu 商品轮播图。"""


@dataclass
class ProcessingConfig:
    llm: LLMConfig
    image_llm: LLMConfig = field(default_factory=lambda: LLMConfig(provider="openai", model="gpt-image-1"))
    template_file: str = ""
    split_output_directory: str = ""
    output_directory: str = ""
    output_filename: str = ""
    title_max_length: int = 100
    filter_clothing: bool = True
    deduplicate: bool = True
    dedupe_field: str = "title"
    enable_scraper: bool = True
    max_concurrent_scrapes: int = 3
    scraper_headless: bool = True
    scraper_proxy: str = ""
    download_images: bool = False
    image_tune_count: int = 2
    image_tune_prompt: str = DEFAULT_IMAGE_TUNE_PROMPT
    description_image_count: int = 4
    default_price: float = 9.99
    default_stock: int = 100
    default_ship_days: str = "2"
    default_length_cm: float = 10
    default_width_cm: float = 10
    default_height_cm: float = 5
    default_weight_g: float = 200
    image_max_px: int = 1000
    enhance_white_bg: bool = False
    request_timeout: int = 20
    retry_count: int = 2
    default_variant_name: str = "颜色"
    default_variant_value: str = "如图"
    default_package_shape: str = "不规则"
    default_package_type: str = "硬包装"


@dataclass
class ProcessingResult:
    output_file: Path
    output_dir: Path
    success_count: int
    failure_count: int
    skipped_count: int
    failures: list[str]


ProgressCallback = Callable[[int, int, str], None]
LogCallback = Callable[[str], None]


def load_config(path: Path) -> ProcessingConfig:
    if not path.exists():
        return ProcessingConfig(llm=LLMConfig())
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    llm_data = data.pop("llm", {})
    image_llm_data = data.pop("image_llm", {})
    valid_fields = set(ProcessingConfig.__dataclass_fields__)
    clean_data = {key: value for key, value in data.items() if key in valid_fields}
    return ProcessingConfig(
        llm=LLMConfig(**llm_data),
        image_llm=LLMConfig(**image_llm_data) if image_llm_data else LLMConfig(provider="openai", model="gpt-image-1"),
        **clean_data,
    )


def save_config(config: ProcessingConfig, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(asdict(config), ensure_ascii=False, indent=2), encoding="utf-8")


def split_excel(
    input_file: Path,
    rows_per_file: int,
    output_dir: Path,
    progress: Optional[ProgressCallback] = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    split_dir = output_dir / f"split_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    split_dir.mkdir(parents=True, exist_ok=True)

    df = read_source_table(input_file)
    total = len(df)
    if total == 0:
        raise ValueError("Excel 文件没有可拆分的数据行")
    if rows_per_file <= 0:
        raise ValueError("每个文件行数必须大于 0")

    batch_count = math.ceil(total / rows_per_file)
    generated: list[Path] = []
    for index in range(batch_count):
        start = index * rows_per_file
        end = min(start + rows_per_file, total)
        batch = df.iloc[start:end]
        file_name = f"第{index + 1:02d}批_{start + 1}-{end}行.xlsx"
        target = split_dir / file_name
        batch.to_excel(target, index=False)
        generated.append(target)
        if progress:
            progress(index + 1, batch_count, f"已生成 {file_name}")

    zip_path = split_dir / "split_files.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_path in generated:
            archive.write(file_path, arcname=file_path.name)
    if progress:
        progress(batch_count, batch_count, f"拆分完成：{zip_path}")
    return zip_path


def read_source_table(file_path: Path) -> pd.DataFrame:
    """Read Excel, real CSV, and cloud exports whose bytes are xlsx but extension is .csv."""

    suffix = file_path.suffix.lower()
    if _is_zip_excel(file_path):
        frame = pd.read_excel(BytesIO(file_path.read_bytes()), header=None, engine="openpyxl")
        return _promote_detected_header(frame)

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        frame = pd.read_excel(file_path, header=None)
        return _promote_detected_header(frame)

    for encoding in ("utf-8-sig", "gb18030", "utf-16", "latin1"):
        try:
            frame = pd.read_csv(file_path, encoding=encoding, header=None)
            return _promote_detected_header(frame)
        except Exception:
            continue
    raise ValueError(f"无法读取文件：{file_path}")


class TemuProcessor:
    def __init__(
        self,
        config: ProcessingConfig,
        progress: Optional[ProgressCallback] = None,
        log: Optional[LogCallback] = None,
    ) -> None:
        self.config = config
        self.optimizer = LLMOptimizer(config.llm)
        self.progress = progress
        self.log = log or (lambda message: None)
        self.seen: set[str] = set()

    def process_files(self, files: Iterable[Path], output_root: Path) -> ProcessingResult:
        output_dir = output_root / f"temu_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        image_dir = output_dir / "images"
        output_dir.mkdir(parents=True, exist_ok=True)
        image_dir.mkdir(parents=True, exist_ok=True)

        rows = self._load_rows(files)
        total = len(rows)
        output_rows: list[dict[str, object]] = []
        failures: list[str] = []
        skipped = 0

        for index, row in enumerate(rows, start=1):
            source_title = self._first(row, ["商品标题（中文）", "商品标题", "产品标题", "标题", "*产品标题", "title", "Title"])
            sku = self._first(row, ["SKU货号", "SKU", "sku", "货号"])
            product_id = self._first(row, ["商品ID", "产品货号", "product_id"])
            category = self._first(row, ["前台分类（英文）", "前台分类（中文）", "后台分类", "类目", "分类", "category", "Category"])
            self._emit(index, total, f"处理中 {index}/{total}: {source_title or sku or '未命名商品'}")

            try:
                if self.config.filter_clothing and self._is_clothing(source_title, category):
                    skipped += 1
                    self.log(f"跳过需尺码表类目：{source_title}")
                    continue

                dedupe_key = self._dedupe_key(source_title, sku)
                if self.config.deduplicate and dedupe_key in self.seen:
                    skipped += 1
                    self.log(f"跳过去重商品：{dedupe_key}")
                    continue
                self.seen.add(dedupe_key)

                existing_english_title = self._first(row, ["商品标题（英文）", "*英文标题", "英文标题"])
                title_seed = existing_english_title or source_title
                english_title = self.optimizer.optimize_title(
                    title_seed,
                    category,
                    self.config.title_max_length,
                    source_title=source_title,
                    sensitive_words_file=self._sensitive_words_file(),
                )
                chinese_title = self.optimizer.translate_title_to_chinese(english_title, source_title, category)
                output_sku = self._listing_sku(sku, product_id, english_title, index)
                image_values = self._process_images(row, image_dir, output_sku or f"item_{index}")
                description = self._append_description_images("", image_values)

                output_rows.append(self._build_output_row(row, chinese_title, english_title, description, output_sku, image_values))
            except Exception as exc:
                failures.append(f"第 {index} 行失败：{exc}")
                self.log(f"失败：第 {index} 行，{exc}")

        output_file = output_dir / self._output_filename()
        self._write_output(output_rows, output_file)
        return ProcessingResult(output_file, output_dir, len(output_rows), len(failures), skipped, failures)

    def _load_rows(self, files: Iterable[Path]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for file_path in files:
            self.log(f"读取文件：{file_path}")
            frame = read_source_table(file_path)
            if self.config.enable_scraper:
                frame = self._run_scraper_enrichment(frame)
            rows.extend(frame.fillna("").to_dict("records"))
        return rows

    def _run_scraper_enrichment(self, frame: pd.DataFrame) -> pd.DataFrame:
        link_column = self._find_product_link_column(frame)
        if not link_column:
            self.log("未找到商品链接列，跳过详情爬虫补全")
            return frame
        urls = [str(value).strip() for value in frame[link_column].fillna("").tolist()]
        if not any(url.startswith(("http://", "https://")) for url in urls):
            self.log("商品链接列为空，跳过详情爬虫补全")
            return frame
        try:
            return asyncio.run(self.enrich_products_with_scraper(frame))
        except RuntimeError:
            loop = asyncio.new_event_loop()
            try:
                return loop.run_until_complete(self.enrich_products_with_scraper(frame))
            finally:
                loop.close()
        except Exception as exc:
            self.log(f"详情爬虫补全失败，继续执行原流程：{exc}")
            return frame

    async def enrich_products_with_scraper(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据补全主方法。爬取失败不阻断主处理流程。"""

        link_column = self._find_product_link_column(df)
        if not link_column:
            return df

        enriched = df.copy()
        scraper_config = ScraperConfig(
            max_concurrent=max(1, int(self.config.max_concurrent_scrapes or 3)),
            headless=bool(self.config.scraper_headless),
            proxy=self.config.scraper_proxy,
            retry_count=max(0, int(self.config.retry_count or 0)),
            max_images=12,
        )
        self.log(f"开始商品详情爬虫补全：{link_column}，并发 {scraper_config.max_concurrent}")

        async with TemuScraper(scraper_config, log=self.log) as scraper:
            tasks = []
            for row_index, value in enriched[link_column].fillna("").items():
                url = str(value).strip()
                if url.startswith(("http://", "https://")):
                    tasks.append((row_index, url, asyncio.create_task(scraper.scrape_product(url))))

            total = len(tasks)
            for current, (row_index, url, task) in enumerate(tasks, start=1):
                try:
                    data = await task
                    if data.get("ok"):
                        self._merge_scraped_data(enriched, row_index, data)
                        self.log(f"爬虫补全成功 {current}/{total}：{url}")
                    else:
                        self.log(f"爬虫补全失败 {current}/{total}：{url} | {data.get('error', '')}")
                except Exception as exc:
                    self.log(f"爬虫补全异常 {current}/{total}：{url} | {exc}")

        return enriched

    def _write_output(self, rows: list[dict[str, object]], output_file: Path) -> None:
        template = Path(self.config.template_file) if self.config.template_file else None
        if template and template.exists():
            workbook = load_workbook(template)
            worksheet = workbook["导入模板"] if "导入模板" in workbook.sheetnames else workbook.active
            headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
            for row_index in range(2, worksheet.max_row + 1):
                for column_index in range(1, worksheet.max_column + 1):
                    worksheet.cell(row_index, column_index).value = None
            for excel_row, item in enumerate(rows, start=2):
                for column_index, header in enumerate(headers, start=1):
                    if header:
                        worksheet.cell(excel_row, column_index).value = item.get(str(header), "")
            for extra_header in ("是否无属性", "变体数量"):
                if extra_header not in headers:
                    column_index = worksheet.max_column + 1
                    worksheet.cell(1, column_index).value = extra_header
                    for excel_row, item in enumerate(rows, start=2):
                        worksheet.cell(excel_row, column_index).value = item.get(extra_header, "")
            workbook.save(output_file)
            return

        pd.DataFrame(rows, columns=DXM_COLUMNS).to_excel(output_file, index=False)

    def _build_output_row(
        self,
        row: dict[str, object],
        source_title: str,
        english_title: str,
        description: str,
        sku: str,
        image_values: dict[str, str],
    ) -> dict[str, object]:
        product_link = self._first(row, ["商品链接", "站外产品链接", "product_url"])
        price = self._first(row, ["*申报价格\n(店铺币种)", "美元价格($)", "价格", "售价", "price"]) or self.config.default_price
        suggestion_price = self._first(row, ["建议售价（USD）", "美元价格($)", "price"]) or price
        preview = image_values.get("preview", self._first(row, ["预览图", "商品主图", "preview_image"]))
        carousel = image_values.get("carousel", self._first(row, ["*轮播图", "商品轮播图", "轮播图"]))
        material = image_values.get("materials", self._first(row, ["*产品素材图", "商品主图", "素材图", "主图"]))
        package_image = image_values.get("package", self._first(row, ["外包装图片", "包装图"]))

        return {
            "*产品标题": source_title,
            "*英文标题": english_title,
            "产品描述": description,
            "产品货号": "",
            "*变种属性名称一": self._variant_name_one(row),
            "*变种属性值一": self._variant_value_one(row),
            "变种属性名称二": self._variant_name_two(row),
            "变种属性值二": self._variant_value_two(row),
            "预览图": preview,
            "*申报价格\n(店铺币种)": price,
            "SKU货号": "",
            "*长（cm）": self._first(row, ["*长（cm）", "长(cm)", "长", "length"]) or self.config.default_length_cm,
            "*宽（cm）": self._first(row, ["*宽（cm）", "宽(cm)", "宽", "width"]) or self.config.default_width_cm,
            "*高（cm）": self._first(row, ["*高（cm）", "高(cm)", "高", "height"]) or self.config.default_height_cm,
            "*重量（g）": self._first(row, ["*重量（g）", "重量(g)", "重量", "weight"]) or self.config.default_weight_g,
            "识别码类型": self._first(row, ["识别码类型"]),
            "识别码": self._first(row, ["识别码"]),
            "站外产品链接": product_link,
            "*轮播图": carousel or preview,
            "*产品素材图": material or preview,
            "外包装形状": self._first(row, ["外包装形状"]) or self.config.default_package_shape,
            "外包装类型": self._first(row, ["外包装类型"]) or self.config.default_package_type,
            "外包装图片": package_image,
            "建议售价（USD）": suggestion_price,
            "库存": self._first(row, ["库存", "stock"]) or self.config.default_stock,
            "发货时效（天）": self._first(row, ["发货时效（天）", "发货时效"]) or self.config.default_ship_days,
            "是否无属性": self._first(row, ["是否无属性", "is_no_attribute"]) or "未知",
            "变体数量": self._first(row, ["变体数量", "variant_count"]) or "",
        }

    def _append_description_images(self, description: str, image_values: dict[str, str]) -> str:
        urls = self._description_image_urls(image_values)
        if not urls:
            return ""

        return "\n".join(f'<p><img src="{escape(url, quote=True)}" /></p>' for url in urls)

    def _description_image_urls(self, image_values: dict[str, str]) -> list[str]:
        count = max(0, int(self.config.description_image_count or 0))
        if count == 0:
            return []

        selected: list[str] = []
        seen: set[str] = set()
        preferred_keys = ("carousel",) if image_values.get("carousel") else ("preview", "materials", "package")
        for key in preferred_keys:
            for value in self._split_image_values(image_values.get(key, "")):
                if value in seen:
                    continue
                selected.append(value)
                seen.add(value)
                if len(selected) >= count:
                    return selected
        return selected

    def _process_images(self, row: dict[str, object], image_dir: Path, sku: str) -> dict[str, str]:
        image_sources = {
            "preview": self._first(row, ["预览图", "商品主图", "preview_image", "图片", "image"]),
            "carousel": self._first(row, ["*轮播图", "商品轮播图", "轮播图"]),
            "materials": self._first(row, ["*产品素材图", "商品主图", "素材图", "主图", "main_image"]),
            "package": self._first(row, ["外包装图片", "包装图", "package_image"]),
        }
        result: dict[str, str] = {}
        for key, value in image_sources.items():
            urls = self._split_image_values(value)
            if not self.config.download_images:
                if urls:
                    result[key] = "\n".join(urls)
                continue
            local_paths = []
            for image_index, url in enumerate(urls, start=1):
                path = self._download_and_optimize_image(url, image_dir, f"{sku}_{key}_{image_index}")
                if path:
                    local_paths.append(str(path))
            if local_paths:
                result[key] = "\n".join(local_paths)
        return result

    def _download_and_optimize_image(self, value: str, image_dir: Path, stem: str) -> str:
        if not value:
            return ""
        parsed = urlparse(value)
        suffix = Path(parsed.path).suffix.lower() or ".jpg"
        suffix = suffix if suffix in {".jpg", ".jpeg", ".png", ".webp"} else ".jpg"
        raw_path = image_dir / f"{self._safe_name(stem)}{suffix}"

        for attempt in range(self.config.retry_count + 1):
            try:
                if parsed.scheme in {"http", "https"}:
                    response = requests.get(value, timeout=self.config.request_timeout)
                    response.raise_for_status()
                    raw_path.write_bytes(response.content)
                else:
                    source = Path(value)
                    if source.exists():
                        shutil.copyfile(source, raw_path)
                    else:
                        return value
                return self._optimize_image(raw_path)
            except Exception as exc:
                if attempt >= self.config.retry_count:
                    self.log(f"图片处理失败：{value}，{exc}")
                else:
                    time.sleep(0.5 * (attempt + 1))
        return value

    def _optimize_image(self, path: Path) -> str:
        with Image.open(path) as image:
            image = image.convert("RGB")
            image.thumbnail((self.config.image_max_px, self.config.image_max_px))
            if self.config.enhance_white_bg:
                image = ImageEnhance.Brightness(image).enhance(1.04)
                image = ImageEnhance.Contrast(image).enhance(1.06)
            target = path.with_suffix(".jpg")
            image.save(target, "JPEG", quality=88, optimize=True)
        if target != path and path.exists():
            path.unlink(missing_ok=True)
        return str(target)

    @staticmethod
    def _find_product_link_column(frame: pd.DataFrame) -> str:
        candidates = [
            "商品链接",
            "站外产品链接",
            "产品链接",
            "详情页",
            "详情页链接",
            "product_url",
            "Product URL",
            "url",
            "URL",
            "link",
            "Link",
        ]
        for name in candidates:
            if name in frame.columns:
                return name
        for column in frame.columns:
            text = str(column)
            lowered = text.lower()
            if "url" in lowered or "link" in lowered or "链接" in text:
                return text
        return ""

    def _merge_scraped_data(self, frame: pd.DataFrame, row_index: int, data: dict[str, object]) -> None:
        title = str(data.get("title") or "").strip()
        description = str(data.get("description") or "").strip()
        images = [str(value).strip() for value in data.get("images", []) if str(value).strip()]
        specs = data.get("specs") if isinstance(data.get("specs"), dict) else {}

        if title:
            self._fill_if_empty(frame, row_index, "商品标题（英文）", title)
            self._set_value(frame, row_index, "爬虫商品标题", title)
        if description:
            self._fill_if_empty(frame, row_index, "产品描述", description)
            self._set_value(frame, row_index, "爬虫产品描述", description)
        if images:
            joined_images = "\n".join(images[:12])
            self._fill_if_empty(frame, row_index, "商品轮播图", joined_images)
            self._fill_if_empty(frame, row_index, "*轮播图", joined_images)
            self._fill_if_empty(frame, row_index, "商品主图", images[0])
            self._fill_if_empty(frame, row_index, "预览图", images[0])
            self._fill_if_empty(frame, row_index, "*产品素材图", joined_images)
        if specs:
            self._set_value(frame, row_index, "爬虫规格参数", specs_to_json(specs))
            self._fill_variant_from_specs(frame, row_index, specs)
        if "is_no_attribute" in data:
            self._set_value(frame, row_index, "是否无属性", "是" if data.get("is_no_attribute") else "否")
        if data.get("variant_count") is not None:
            self._set_value(frame, row_index, "变体数量", data.get("variant_count"))
        detected_attributes = data.get("detected_attributes")
        if detected_attributes:
            self._set_value(frame, row_index, "检测到的变种属性", ", ".join(str(value) for value in detected_attributes))

        self._fill_if_empty(frame, row_index, "商品ID", str(data.get("product_id") or ""))
        self._fill_if_empty(frame, row_index, "SKU", str(data.get("sku") or ""))
        self._fill_numeric_if_present(frame, row_index, "重量(g)", data.get("weight_g"))
        self._fill_numeric_if_present(frame, row_index, "长(cm)", data.get("length_cm"))
        self._fill_numeric_if_present(frame, row_index, "宽(cm)", data.get("width_cm"))
        self._fill_numeric_if_present(frame, row_index, "高(cm)", data.get("height_cm"))

    def _fill_variant_from_specs(self, frame: pd.DataFrame, row_index: int, specs: dict[str, object]) -> None:
        lowered = {str(key).strip().lower(): str(value).strip() for key, value in specs.items() if str(value).strip()}
        color = self._first_matching_spec(lowered, ["color", "colour", "颜色", "颜色分类"])
        size = self._first_matching_spec(lowered, ["size", "尺寸", "尺码", "规格"])
        material = self._first_matching_spec(lowered, ["material", "材质", "fabric", "composition"])

        if color:
            self._fill_if_empty(frame, row_index, "颜色", color)
            self._fill_if_empty(frame, row_index, "变种属性名称一", "颜色")
            self._fill_if_empty(frame, row_index, "变种属性值一", color)
            if size:
                self._fill_if_empty(frame, row_index, "尺寸", size)
                self._fill_if_empty(frame, row_index, "变种属性名称二", "尺寸")
                self._fill_if_empty(frame, row_index, "变种属性值二", size)
        elif size:
            self._fill_if_empty(frame, row_index, "规格", size)
            self._fill_if_empty(frame, row_index, "变种属性名称一", "尺寸")
            self._fill_if_empty(frame, row_index, "变种属性值一", size)
        if material:
            self._fill_if_empty(frame, row_index, "材质", material)

    @staticmethod
    def _first_matching_spec(specs: dict[str, str], keys: list[str]) -> str:
        for key in keys:
            key_lower = key.lower()
            for spec_key, value in specs.items():
                if key_lower == spec_key or key_lower in spec_key:
                    return value
        return ""

    @staticmethod
    def _set_value(frame: pd.DataFrame, row_index: int, column: str, value: object) -> None:
        if column not in frame.columns:
            frame[column] = ""
        frame.at[row_index, column] = value

    def _fill_if_empty(self, frame: pd.DataFrame, row_index: int, column: str, value: object) -> None:
        if value is None or str(value).strip() == "":
            return
        if column not in frame.columns:
            frame[column] = ""
        current = frame.at[row_index, column]
        if pd.isna(current) or str(current).strip() == "":
            frame.at[row_index, column] = value

    def _fill_numeric_if_present(self, frame: pd.DataFrame, row_index: int, column: str, value: object) -> None:
        if value is None or str(value).strip() == "":
            return
        self._fill_if_empty(frame, row_index, column, value)

    def _emit(self, current: int, total: int, message: str) -> None:
        self.log(message)
        if self.progress:
            self.progress(current, total, message)

    @staticmethod
    def _sensitive_words_file() -> Path:
        return Path(__file__).resolve().parent / "config" / "sensitive_words.json"

    def _dedupe_key(self, title: str, sku: str) -> str:
        value = sku if self.config.dedupe_field.lower() == "sku" else title
        return re.sub(r"\s+", " ", str(value or "")).strip().lower()

    def _variant_name_one(self, row: dict[str, object]) -> str:
        configured = self._first(row, ["*变种属性名称一", "变种属性名称一", "属性名称一", "规格名称", "规格名", "变体名称", "变种名称"])
        if configured:
            return configured
        if self._color_value(row):
            return "颜色"
        if self._size_value(row):
            return "尺寸"
        if self._material_value(row):
            return "材质"
        return self.config.default_variant_name or "颜色"

    def _variant_value_one(self, row: dict[str, object]) -> str:
        value = self._first(
            row,
            [
                "*变种属性值一",
                "变种属性值一",
                "属性值一",
                "颜色",
                "颜色分类",
                "规格",
                "规格值",
                "款式",
                "型号",
                "变体值",
                "变种值",
            ],
        )
        if value:
            return value
        configured = str(self.config.default_variant_value or "").strip()
        if not configured or configured.lower() == "default":
            return "如图"
        return configured

    def _variant_name_two(self, row: dict[str, object]) -> str:
        configured = self._first(row, ["变种属性名称二", "属性名称二", "规格名称二", "规格名二", "变体名称二", "变种名称二"])
        if configured:
            return configured
        if self._color_value(row) and self._size_value(row):
            return "尺寸"
        if self._color_value(row) and self._material_value(row):
            return "材质"
        return ""

    def _variant_value_two(self, row: dict[str, object]) -> str:
        configured = self._first(row, ["变种属性值二", "属性值二", "规格值二", "变体值二", "变种值二"])
        if configured:
            return configured
        if self._color_value(row) and self._size_value(row):
            return self._size_value(row)
        if self._color_value(row) and self._material_value(row):
            return self._material_value(row)
        return ""

    def _color_value(self, row: dict[str, object]) -> str:
        return self._first(row, ["颜色", "颜色分类", "Color", "Colour", "color", "colour"])

    def _size_value(self, row: dict[str, object]) -> str:
        return self._first(row, ["尺寸", "尺码", "规格", "规格值", "Size", "size"])

    def _material_value(self, row: dict[str, object]) -> str:
        return self._first(row, ["材质", "面料", "材料", "Material", "material", "fabric", "composition"])

    def _listing_sku(self, source_sku: str, product_id: str, english_title: str, row_index: int) -> str:
        if source_sku:
            return source_sku
        digits = re.sub(r"\D+", "", str(product_id or ""))
        if digits:
            return f"TAO-{digits[-6:]}-{row_index:03d}"
        return f"{self._make_sku(english_title)}-{row_index:03d}"

    @staticmethod
    def _first(row: dict[str, object], names: list[str]) -> str:
        for name in names:
            value = row.get(name)
            if value is not None and str(value).strip():
                return str(value).strip()
        return ""

    @staticmethod
    def _split_image_values(value: str) -> list[str]:
        if not value:
            return []
        text = str(value).strip().strip("[]")
        return [part.strip().strip("'\"") for part in re.split(r"[;,\n]", text) if part.strip()]

    @staticmethod
    def _is_clothing(title: str, category: str) -> bool:
        text = f"{title} {category}".lower()
        return any(keyword.lower() in text for keyword in CLOTHING_KEYWORDS)

    @staticmethod
    def _make_sku(title: str) -> str:
        base = re.sub(r"[^A-Za-z0-9]+", "-", title.upper()).strip("-")
        return base[:32] or f"TEMU-{int(time.time())}"

    @staticmethod
    def _safe_name(value: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", value)[:80] or "image"

    def _output_filename(self) -> str:
        raw = (self.config.output_filename or "").strip()
        if not raw:
            return "店小秘_TEMU半托管导入模板.xlsx"
        safe = re.sub(r'[<>:"/\\|?*]+', "_", raw).strip().strip(".")
        if not safe:
            return "店小秘_TEMU半托管导入模板.xlsx"
        if not safe.lower().endswith(".xlsx"):
            safe += ".xlsx"
        return safe


def _is_zip_excel(file_path: Path) -> bool:
    with file_path.open("rb") as handle:
        return handle.read(2) == b"PK"


def _promote_detected_header(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.dropna(how="all").reset_index(drop=True)
    header_index = 0
    markers = {"商品标题（中文）", "商品标题（英文）", "商品ID", "*产品标题", "*英文标题"}
    for index, row in frame.head(10).iterrows():
        values = {str(value).strip() for value in row.tolist() if pd.notna(value) and str(value).strip()}
        if values & markers:
            header_index = index
            break
    header = [
        str(value).strip() if pd.notna(value) and str(value).strip() else f"未命名列{idx + 1}"
        for idx, value in enumerate(frame.iloc[header_index].tolist())
    ]
    data = frame.iloc[header_index + 1 :].copy()
    data.columns = header
    return data.dropna(how="all").reset_index(drop=True)
