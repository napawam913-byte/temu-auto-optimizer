from __future__ import annotations

import html
import json
import random
import re
import shutil
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

import pandas as pd
from openpyxl import load_workbook

from image_tuner import ImageTuneError, tune_image
from processor import load_config


def bulk_tune_excel_images(excel_file: Path, config, progress=None, log=None) -> dict[str, object]:
    frame = _read_frame(excel_file).fillna("")
    total = len(frame)
    tune_dir = excel_file.parent / "ai_tuned_images"
    backup = excel_file.with_suffix(".before_image_tune.xlsx")
    if not backup.exists():
        shutil.copyfile(excel_file, backup)

    count = max(1, int(config.image_tune_count or 2))
    product_count = 0
    image_count = 0
    skipped_count = 0
    failures: list[str] = []

    for row_index, row in frame.iterrows():
        if progress:
            progress(row_index + 1, total, f"正在改图 {row_index + 1}/{total}")
        try:
            carousel = _image_values(_first(row, ["*轮播图", "商品轮播图", "*产品素材图"]))
            candidates = [idx for idx, image in enumerate(carousel) if image and not _is_tuned_image(image)]
            if not candidates:
                skipped_count += 1
                continue

            selected_indices = sorted(random.sample(candidates, min(count, len(candidates))))
            row_updated = False
            for image_index in selected_indices:
                original = carousel[image_index]
                tuned_path = tune_image(
                    original,
                    tune_dir,
                    config,
                    f"row_{row_index + 2}_image_{image_index + 1}",
                )
                carousel[image_index] = str(tuned_path)
                row_updated = True
                image_count += 1
                if log:
                    log(f"第 {row_index + 2} 行图片 {image_index + 1} 已微调：{tuned_path}")

            if row_updated:
                _update_carousel_cell(excel_file, row_index, "\n".join(carousel))
                product_count += 1
        except Exception as exc:
            failures.append(f"第 {row_index + 2} 行改图失败：{exc}")
            if log:
                log(failures[-1])

    return {
        "products": product_count,
        "images": image_count,
        "skipped": skipped_count,
        "failures": failures,
        "backup": str(backup),
    }


class PreviewServer:
    def __init__(self, config_path: Path, host: str = "127.0.0.1", port: int = 8765) -> None:
        self.config_path = config_path
        self.host = host
        self.port = port
        self.httpd: ThreadingHTTPServer | None = None
        self.thread: threading.Thread | None = None

    def start(self) -> str:
        if self.httpd:
            return f"http://{self.host}:{self.port}"

        handler = self._make_handler()
        try:
            self.httpd = ThreadingHTTPServer((self.host, self.port), handler)
        except OSError:
            self.httpd = ThreadingHTTPServer((self.host, 0), handler)
            self.port = int(self.httpd.server_address[1])

        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()
        return f"http://{self.host}:{self.port}"

    def _make_handler(self):
        config_path = self.config_path

        class Handler(BaseHTTPRequestHandler):
            def log_message(self, format, *args):  # noqa: A002
                return

            def do_GET(self):  # noqa: N802
                parsed = urlparse(self.path)
                if parsed.path == "/preview":
                    self._handle_preview(parsed)
                elif parsed.path == "/local-image":
                    self._handle_local_image(parsed)
                else:
                    self._send_text("Not found", status=404)

            def do_POST(self):  # noqa: N802
                parsed = urlparse(self.path)
                if parsed.path == "/api/tune":
                    self._handle_tune()
                else:
                    self._send_json({"ok": False, "error": "Not found"}, status=404)

            def _handle_preview(self, parsed) -> None:
                query = parse_qs(parsed.query)
                file_value = query.get("file", [""])[0]
                if not file_value:
                    self._send_text("Missing file", status=400)
                    return
                excel_file = Path(unquote(file_value))
                if not excel_file.exists():
                    self._send_text(f"File not found: {excel_file}", status=404)
                    return
                frame = _read_frame(excel_file).fillna("")
                self._send_html(_render_preview_page(excel_file, frame))

            def _handle_local_image(self, parsed) -> None:
                query = parse_qs(parsed.query)
                path_value = query.get("path", [""])[0]
                image_path = Path(unquote(path_value))
                if not image_path.exists() or not image_path.is_file():
                    self._send_text("Image not found", status=404)
                    return
                suffix = image_path.suffix.lower()
                content_type = "image/jpeg"
                if suffix == ".png":
                    content_type = "image/png"
                elif suffix == ".webp":
                    content_type = "image/webp"
                self.send_response(200)
                self.send_header("Content-Type", content_type)
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(image_path.read_bytes())

            def _handle_tune(self) -> None:
                try:
                    length = int(self.headers.get("Content-Length", "0"))
                    payload = json.loads(self.rfile.read(length).decode("utf-8"))
                    excel_file = Path(payload["file"])
                    row_index = int(payload["rowIndex"])
                    selected_indices = [int(value) for value in payload.get("selectedIndices", [])]
                    mode = payload.get("mode", "selected")

                    config = load_config(config_path)
                    frame = _read_frame(excel_file).fillna("")
                    row = frame.iloc[row_index]
                    carousel = _image_values(_first(row, ["*轮播图", "商品轮播图", "*产品素材图"]))
                    if not carousel:
                        raise ImageTuneError("该商品没有可微调的轮播图")
                    if mode == "random" or not selected_indices:
                        count = max(1, int(config.image_tune_count or 2))
                        selected_indices = sorted(random.sample(range(len(carousel)), min(count, len(carousel))))

                    tune_dir = excel_file.parent / "ai_tuned_images"
                    backup = excel_file.with_suffix(".before_image_tune.xlsx")
                    if not backup.exists():
                        shutil.copyfile(excel_file, backup)

                    updates = []
                    for image_index in selected_indices:
                        if image_index < 0 or image_index >= len(carousel):
                            continue
                        original = carousel[image_index]
                        tuned_path = tune_image(
                            original,
                            tune_dir,
                            config,
                            f"row_{row_index + 2}_image_{image_index + 1}",
                        )
                        carousel[image_index] = str(tuned_path)
                        updates.append(
                            {
                                "index": image_index,
                                "path": str(tuned_path),
                                "src": _image_src(str(tuned_path)),
                            }
                        )
                    _update_carousel_cell(excel_file, row_index, "\n".join(carousel))
                    self._send_json({"ok": True, "updates": updates, "backup": str(backup)})
                except Exception as exc:
                    self._send_json({"ok": False, "error": str(exc)}, status=500)

            def _send_html(self, body: str) -> None:
                encoded = body.encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(encoded)))
                self.end_headers()
                self.wfile.write(encoded)

            def _send_text(self, body: str, status: int = 200) -> None:
                encoded = body.encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "text/plain; charset=utf-8")
                self.send_header("Content-Length", str(len(encoded)))
                self.end_headers()
                self.wfile.write(encoded)

            def _send_json(self, body: dict, status: int = 200) -> None:
                encoded = json.dumps(body, ensure_ascii=False).encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(encoded)))
                self.end_headers()
                self.wfile.write(encoded)

        return Handler


def _read_frame(excel_file: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(excel_file, sheet_name="导入模板")
    except Exception:
        return pd.read_excel(excel_file)


def _render_preview_page(excel_file: Path, frame: pd.DataFrame) -> str:
    cards = []
    for index, row in frame.iterrows():
        title = _first(row, ["*产品标题", "产品标题", "商品标题（中文）"])
        english_title = _first(row, ["*英文标题", "英文标题", "商品标题（英文）"])
        price = _first(row, ["*申报价格\n(店铺币种)", "美元价格($)", "价格"])
        preview_images = _image_values(_first(row, ["预览图", "商品主图"]))
        carousel_images = _image_values(_first(row, ["*轮播图", "商品轮播图", "*产品素材图"]))
        is_modified = any(_is_tuned_image(image) for image in carousel_images)
        card_class = "product-card modified" if is_modified else "product-card"
        cards.append(
            f"""
            <article class="{card_class}" data-row="{index}">
              <header>
                <span class="row-index">#{index + 2}</span>
                <div>
                  <h2>{_esc(english_title or title)}</h2>
                  <p>{_esc(title)}</p>
                </div>
                <strong>{_esc(price)}</strong>
              </header>
              <section class="image-grid">
                <div>
                  <h3>预览图</h3>
                  {_render_images(preview_images[:1], selectable=False)}
                </div>
                <div>
                  <h3>轮播图</h3>
                  {_render_images(carousel_images, selectable=True)}
                  <div class="actions">
                    <button data-action="selected">微调选中图片</button>
                    <button data-action="random">随机微调配置数量</button>
                    <span class="status"></span>
                  </div>
                </div>
              </section>
            </article>
            """
        )

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Temu 上架网页预览</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, "Microsoft YaHei", sans-serif;
      background: #f6f7f9;
      color: #202733;
    }}
    .top {{
      position: sticky;
      top: 0;
      z-index: 10;
      background: rgba(255,255,255,.97);
      border-bottom: 1px solid #d9dee7;
      padding: 14px 22px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      flex-wrap: wrap;
    }}
    .top-actions {{
      display: flex;
      gap: 10px;
      align-items: center;
      flex-wrap: wrap;
    }}
    .pager {{
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
      width: 100%;
      margin-top: 8px;
      color: #667085;
      font-size: 13px;
    }}
    h1 {{ font-size: 18px; margin: 0 0 4px; }}
    .path {{ color: #667085; font-size: 13px; word-break: break-all; }}
    #search {{
      width: min(420px, 36vw);
      padding: 9px 11px;
      border: 1px solid #d9dee7;
      border-radius: 6px;
      font-size: 14px;
    }}
    #page-size {{
      width: 84px;
      padding: 7px 9px;
      border: 1px solid #d9dee7;
      border-radius: 6px;
    }}
    main {{ padding: 18px 22px 42px; display: grid; gap: 14px; }}
    .product-card {{
      background: #fff;
      border: 1px solid #d9dee7;
      border-radius: 8px;
      padding: 14px;
    }}
    .product-card.modified {{
      background: #f0fdf4;
      border-color: #22c55e;
    }}
    .product-card header {{
      display: grid;
      grid-template-columns: auto 1fr auto;
      gap: 12px;
      border-bottom: 1px solid #d9dee7;
      padding-bottom: 12px;
      margin-bottom: 12px;
    }}
    .row-index {{ background: #eef4ff; color: #2563eb; padding: 5px 8px; border-radius: 4px; font-weight: 700; }}
    h2 {{ margin: 0; font-size: 16px; line-height: 1.35; }}
    h3 {{ margin: 0 0 8px; color: #667085; font-size: 13px; }}
    p {{ margin: 5px 0 0; color: #667085; line-height: 1.45; }}
    strong {{ color: #0f766e; white-space: nowrap; }}
    .image-grid {{ display: grid; grid-template-columns: minmax(180px, 220px) minmax(0, 1fr); gap: 16px; }}
    .carousel-strip {{
      display: grid;
      grid-template-columns: auto minmax(0, 1fr) auto;
      gap: 8px;
      align-items: center;
    }}
    .images {{
      display: flex;
      gap: 10px;
      overflow-x: auto;
      scroll-snap-type: x proximity;
      padding: 4px 0 8px;
      scrollbar-width: thin;
    }}
    .scroll-button {{
      min-width: 34px;
      height: 34px;
      padding: 0;
      border-radius: 50%;
      border-color: #d9dee7;
      background: #fff;
      color: #2563eb;
    }}
    .image-cell {{ width: clamp(132px, 16vw, 172px); flex: 0 0 auto; position: relative; scroll-snap-align: start; }}
    .image-cell.tuned {{
      background: #ecfdf3;
      border: 2px solid #22c55e;
      border-radius: 8px;
      padding: 4px;
    }}
    .image-cell img {{
      width: 100%;
      aspect-ratio: 1 / 1;
      object-fit: contain;
      border: 1px solid #d9dee7;
      border-radius: 6px;
      background: #fff;
    }}
    .image-cell label {{
      display: flex;
      gap: 6px;
      align-items: center;
      margin-top: 5px;
      font-size: 12px;
    }}
    .url {{ margin-top: 4px; color: #667085; font-size: 11px; overflow-wrap: anywhere; }}
    .badge {{
      display: none;
      margin-top: 5px;
      color: #166534;
      background: #dcfce7;
      border-radius: 4px;
      padding: 3px 6px;
      font-size: 12px;
      font-weight: 700;
    }}
    .image-cell.tuned .badge {{ display: inline-block; }}
    .actions {{ display: flex; align-items: center; gap: 8px; margin-top: 8px; }}
    button {{
      border: 1px solid #2563eb;
      background: #2563eb;
      color: #fff;
      border-radius: 6px;
      padding: 8px 10px;
      cursor: pointer;
    }}
    button + button {{ background: #fff; color: #2563eb; }}
    .pager button {{
      background: #fff;
      color: #2563eb;
      padding: 6px 9px;
    }}
    #tune-unmodified {{ background: #16a34a; border-color: #16a34a; }}
    button:disabled {{ opacity: .55; cursor: wait; }}
    .status {{ color: #667085; font-size: 13px; }}
    .empty {{ border: 1px dashed #d9dee7; color: #667085; padding: 18px; border-radius: 6px; }}
    @media (max-width: 860px) {{
      .top {{ flex-direction: column; align-items: stretch; }}
      .top-actions {{ align-items: stretch; }}
      #search {{ width: 100%; }}
      .product-card header, .image-grid {{ grid-template-columns: 1fr; }}
      .scroll-button {{ display: none; }}
      .image-cell {{ width: 138px; }}
    }}
  </style>
</head>
<body>
  <header class="top">
    <div>
      <h1>Temu 上架网页预览</h1>
      <div class="path">{_esc(str(excel_file))} · 共 {len(frame)} 行</div>
    </div>
    <div class="top-actions">
      <input id="search" placeholder="搜索标题、英文标题或价格">
      <button id="tune-unmodified">一键微调未修改商品</button>
      <span id="global-status" class="status"></span>
    </div>
    <div class="pager">
      <label>每页商品数 <input id="page-size" type="number" min="1" max="500" value="20"></label>
      <button id="prev-page">上一页</button>
      <button id="next-page">下一页</button>
      <span id="page-info"></span>
    </div>
  </header>
  <main>{''.join(cards)}</main>
  <script>
    const excelFile = {json.dumps(str(excel_file), ensure_ascii=False)};
    const cards = Array.from(document.querySelectorAll('.product-card'));
    const search = document.getElementById('search');
    const pageSizeInput = document.getElementById('page-size');
    const pageInfo = document.getElementById('page-info');
    let currentPage = 1;
    let filteredCards = cards;
    function applyPagination() {{
      const pageSize = Math.max(1, Number(pageSizeInput.value) || 20);
      const totalPages = Math.max(1, Math.ceil(filteredCards.length / pageSize));
      currentPage = Math.min(Math.max(1, currentPage), totalPages);
      const start = (currentPage - 1) * pageSize;
      const visible = new Set(filteredCards.slice(start, start + pageSize));
      for (const card of cards) {{
        card.style.display = visible.has(card) ? '' : 'none';
      }}
      pageInfo.textContent = `第 ${{currentPage}} / ${{totalPages}} 页，共 ${{filteredCards.length}} 个商品`;
      document.getElementById('prev-page').disabled = currentPage <= 1;
      document.getElementById('next-page').disabled = currentPage >= totalPages;
    }}
    function applyFilter() {{
      const term = search.value.trim().toLowerCase();
      filteredCards = cards.filter(card => card.innerText.toLowerCase().includes(term));
      currentPage = 1;
      applyPagination();
    }}
    search.addEventListener('input', applyFilter);
    pageSizeInput.addEventListener('change', () => {{
      currentPage = 1;
      applyPagination();
    }});
    document.getElementById('prev-page').addEventListener('click', () => {{
      currentPage -= 1;
      applyPagination();
    }});
    document.getElementById('next-page').addEventListener('click', () => {{
      currentPage += 1;
      applyPagination();
    }});
    applyFilter();
    document.addEventListener('click', event => {{
      const button = event.target.closest('button[data-scroll]');
      if (!button) return;
      const strip = button.closest('.carousel-strip');
      const images = strip.querySelector('.images');
      const direction = button.dataset.scroll === 'left' ? -1 : 1;
      images.scrollBy({{ left: direction * Math.max(260, images.clientWidth * 0.8), behavior: 'smooth' }});
    }});
    document.getElementById('tune-unmodified').addEventListener('click', async () => {{
      const globalStatus = document.getElementById('global-status');
      const pending = cards.filter(card => !card.classList.contains('modified') && card.querySelector('.image-cell'));
      if (!pending.length) {{
        globalStatus.textContent = '没有未修改商品';
        return;
      }}
      const button = document.getElementById('tune-unmodified');
      button.disabled = true;
      let done = 0;
      for (const card of pending) {{
        globalStatus.textContent = `正在微调 ${{done + 1}} / ${{pending.length}}`;
        await tuneCard(card, 'random', []);
        done += 1;
      }}
      globalStatus.textContent = `一键微调完成 ${{done}} 个商品`;
      button.disabled = false;
    }});
    document.addEventListener('click', async event => {{
      const button = event.target.closest('button[data-action]');
      if (!button) return;
      const card = button.closest('.product-card');
      const status = card.querySelector('.status');
      const selected = Array.from(card.querySelectorAll('input[type="checkbox"]:checked')).map(input => Number(input.dataset.index));
      const mode = button.dataset.action === 'random' ? 'random' : 'selected';
      if (mode === 'selected' && selected.length === 0) {{
        status.textContent = '请先勾选要微调的轮播图';
        return;
      }}
      status.textContent = '正在调用图片模型微调...';
      await tuneCard(card, mode, selected);
    }});
    async function tuneCard(card, mode, selected) {{
      const status = card.querySelector('.status');
      const buttons = Array.from(card.querySelectorAll('button'));
      buttons.forEach(item => item.disabled = true);
      try {{
        const response = await fetch('/api/tune', {{
          method: 'POST',
          headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{
            file: excelFile,
            rowIndex: Number(card.dataset.row),
            selectedIndices: selected,
            mode
          }})
        }});
        const result = await response.json();
        if (!result.ok) throw new Error(result.error || '微调失败');
        for (const update of result.updates) {{
          const cell = card.querySelector(`.image-cell[data-index="${{update.index}}"]`);
          if (cell) {{
            card.classList.add('modified');
            cell.classList.add('tuned');
            cell.querySelector('img').src = update.src + '&t=' + Date.now();
            cell.querySelector('.url').textContent = update.path;
            cell.querySelector('input').checked = false;
          }}
        }}
        status.textContent = `完成 ${{result.updates.length}} 张，已更新 Excel`;
      }} catch (error) {{
        status.textContent = error.message;
      }} finally {{
        buttons.forEach(item => item.disabled = false);
      }}
    }}
  </script>
</body>
</html>"""


def _render_images(images: list[str], selectable: bool) -> str:
    if not images:
        return '<div class="empty">暂无图片</div>'
    cells = []
    for index, image in enumerate(images):
        safe_image = _esc(image)
        tuned_class = " tuned" if _is_tuned_image(image) else ""
        checkbox = (
            f'<label><input type="checkbox" data-index="{index}"> 选择图 {index + 1}</label>'
            if selectable
            else ""
        )
        cells.append(
            f"""
            <div class="image-cell{tuned_class}" data-index="{index}">
              <img loading="lazy" src="{_esc(_image_src(image))}" alt="product image">
              {checkbox}
              <div class="badge">已修改</div>
              <div class="url">{safe_image}</div>
            </div>
            """
        )
    return (
        '<div class="carousel-strip">'
        '<button class="scroll-button" data-scroll="left" title="向左滚动">‹</button>'
        f'<div class="images">{"".join(cells)}</div>'
        '<button class="scroll-button" data-scroll="right" title="向右滚动">›</button>'
        '</div>'
    )


def _update_carousel_cell(excel_file: Path, row_index: int, value: str) -> None:
    workbook = load_workbook(excel_file)
    worksheet = workbook["导入模板"] if "导入模板" in workbook.sheetnames else workbook.active
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    try:
        column_index = headers.index("*轮播图") + 1
    except ValueError:
        raise ImageTuneError("Excel 中没有找到 *轮播图 列")
    worksheet.cell(row_index + 2, column_index).value = value
    try:
        workbook.save(excel_file)
    except PermissionError as exc:
        if _update_open_spreadsheet_app(excel_file, row_index + 2, column_index, value):
            return
        raise ImageTuneError(
            "结果 Excel 正在被占用，且无法通过已打开的 Excel/WPS 写入。"
            f"请关闭文件后重试，或安装 pywin32 后用 Microsoft Excel 打开：{excel_file}"
        ) from exc
    finally:
        workbook.close()


def _update_open_spreadsheet_app(excel_file: Path, excel_row: int, column_index: int, value: str) -> bool:
    """Write through an already-open Excel/WPS workbook when Windows locks the xlsx file."""

    try:
        import win32com.client  # type: ignore
    except Exception:
        return False

    target = str(excel_file.resolve()).lower()
    for app_name in ("Excel.Application", "Ket.Application"):
        try:
            app = win32com.client.GetActiveObject(app_name)
        except Exception:
            continue
        try:
            workbooks = app.Workbooks
            for index in range(1, int(workbooks.Count) + 1):
                workbook = workbooks.Item(index)
                try:
                    full_name = str(workbook.FullName).lower()
                except Exception:
                    continue
                if full_name != target:
                    continue
                try:
                    worksheet = workbook.Worksheets("导入模板")
                except Exception:
                    worksheet = workbook.ActiveSheet
                worksheet.Cells(excel_row, column_index).Value = value
                workbook.Save()
                return True
        except Exception:
            continue
    return False


def _first(row, names: list[str]) -> str:
    for name in names:
        if name in row and str(row[name]).strip():
            return str(row[name]).strip()
    return ""


def _image_values(value: str) -> list[str]:
    text = str(value or "").strip().strip("[]")
    if not text:
        return []
    return [part.strip().strip("'\"") for part in re.split(r"[;,\n\r]+", text) if part.strip()]


def _image_src(value: str) -> str:
    if value.startswith(("http://", "https://")):
        return value
    return "/local-image?path=" + quote(value)


def _is_tuned_image(value: str) -> bool:
    normalized = str(value or "").replace("\\", "/").lower()
    return "/ai_tuned_images/" in normalized or normalized.startswith("ai_tuned_images/")


def _esc(value: object) -> str:
    return html.escape(str(value or ""), quote=True)
